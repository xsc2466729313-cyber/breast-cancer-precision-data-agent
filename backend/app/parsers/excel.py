from __future__ import annotations

import io

from openpyxl import load_workbook

from backend.app.parsers.base import ParseRequest, ParseResult, ParsedRecord


class ExcelParser:
    version = "excel-openpyxl-v1"

    def parse(self, request: ParseRequest, *, workbook_bytes: bytes | None = None) -> ParseResult:
        if workbook_bytes is None and request.rows:
            records = []
            for index, row in enumerate(request.rows):
                for field, value in row.items():
                    records.append(
                        ParsedRecord(
                            source_id=request.source_id,
                            source_file=request.filename,
                            source_location=f"sheet:inline:row:{index + 2}",
                            raw_field=str(field),
                            raw_value=value,
                            inferred_semantic_type="categorical" if isinstance(value, str) else "continuous",
                            parse_method=self.version,
                            parse_confidence=0.85,
                            parser_version=self.version,
                            status="PARSED",
                        )
                    )
            return ParseResult(source_id=request.source_id, parse_method=self.version, records=records, status="PARSED" if records else "FAILED")
        if not workbook_bytes:
            return ParseResult(
                source_id=request.source_id,
                parse_method=self.version,
                records=[],
                warnings=["未提供 Excel 字节；不要让模型猜测表格。"],
                status="REVIEW",
            )
        workbook = load_workbook(io.BytesIO(workbook_bytes), data_only=True)
        records: list[ParsedRecord] = []
        warnings: list[str] = []
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            header = [str(item).strip() if item is not None else f"column_{index}" for index, item in enumerate(rows[0])]
            for row_index, row in enumerate(rows[1:], start=2):
                for col_index, value in enumerate(row):
                    records.append(
                        ParsedRecord(
                            source_id=request.source_id,
                            source_file=request.filename,
                            source_location=f"sheet:{sheet.title}:row:{row_index}:col:{col_index + 1}",
                            raw_field=header[col_index] if col_index < len(header) else f"column_{col_index}",
                            raw_value=value,
                            inferred_semantic_type="missing" if value in {None, ""} else type(value).__name__,
                            parse_method=self.version,
                            parse_confidence=0.8 if value not in {None, ""} else 0.35,
                            parser_version=self.version,
                            status="PARSED",
                        )
                    )
            if any(sheet.merged_cells.ranges):
                warnings.append(f"工作表 {sheet.title} 含合并单元格，需要人工核对表头。")
        status = "REVIEW" if warnings else ("PARSED" if records else "FAILED")
        return ParseResult(source_id=request.source_id, parse_method=self.version, records=records, warnings=warnings, status=status)
