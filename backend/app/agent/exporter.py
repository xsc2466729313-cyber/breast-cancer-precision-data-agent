from __future__ import annotations

import csv
import io
import json
from dataclasses import dataclass
from enum import Enum
from typing import Any

import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from backend.app.agent.models import AgentTaskResult


class AgentExportFormat(str, Enum):
    CSV = "csv"
    PARQUET = "parquet"
    XLSX = "xlsx"
    JSON = "json"
    METADATA = "metadata"
    QUALITY_REPORT = "quality_report"


@dataclass(frozen=True)
class AgentExportedDataset:
    content: bytes
    media_type: str
    filename: str


class AgentDatasetExportService:
    def export(
        self,
        result: AgentTaskResult,
        file_format: AgentExportFormat,
    ) -> AgentExportedDataset:
        dataset = result.modeling_dataset
        if file_format in {AgentExportFormat.CSV, AgentExportFormat.PARQUET, AgentExportFormat.XLSX, AgentExportFormat.JSON} and not dataset.rows:
            raise ValueError("当前任务没有可导出的科研数据行。")
        if file_format == AgentExportFormat.CSV:
            content = self._csv(result)
            media_type = "text/csv; charset=utf-8"
            filename = f"{result.task_id}-科研数据集.csv"
        elif file_format == AgentExportFormat.PARQUET:
            content = self._parquet(result)
            media_type = "application/vnd.apache.parquet"
            filename = f"{result.task_id}-科研数据集.parquet"
        elif file_format == AgentExportFormat.JSON:
            content = self._json_dataset(result)
            media_type = "application/json; charset=utf-8"
            filename = f"{result.task_id}-科研数据集.json"
        elif file_format == AgentExportFormat.METADATA:
            content = self._metadata(result)
            media_type = "application/json; charset=utf-8"
            filename = f"{result.task_id}-元数据.json"
        elif file_format == AgentExportFormat.QUALITY_REPORT:
            content = self._quality_report(result)
            media_type = "application/json; charset=utf-8"
            filename = f"{result.task_id}-质量报告.json"
        else:
            content = self._xlsx(result)
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename = f"{result.task_id}-科研数据集.xlsx"
        return AgentExportedDataset(
            content=content,
            media_type=media_type,
            filename=filename,
        )

    @staticmethod
    def _csv(result: AgentTaskResult) -> bytes:
        columns = [column.name for column in result.modeling_dataset.columns]
        buffer = io.StringIO(newline="")
        writer = csv.DictWriter(buffer, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for row in result.modeling_dataset.rows:
            writer.writerow({name: AgentDatasetExportService._scalar(row.get(name)) for name in columns})
        return ("\ufeff" + buffer.getvalue()).encode("utf-8")

    @staticmethod
    def _parquet(result: AgentTaskResult) -> bytes:
        arrays: dict[str, pa.Array] = {}
        for column in result.modeling_dataset.columns:
            values = [
                AgentDatasetExportService._scalar(row.get(column.name))
                for row in result.modeling_dataset.rows
            ]
            if column.data_type == "number":
                numeric = [value for value in values if value is not None]
                use_float = any(isinstance(value, float) for value in numeric)
                arrow_type = pa.float64() if use_float else pa.int64()
                normalized = [
                    (float(value) if use_float else int(value)) if value is not None else None
                    for value in values
                ]
            else:
                arrow_type = pa.string()
                normalized = [str(value) if value is not None else None for value in values]
            arrays[column.name] = pa.array(normalized, type=arrow_type)
        table = pa.table(arrays)
        sink = pa.BufferOutputStream()
        pq.write_table(table, sink, compression="snappy")
        return sink.getvalue().to_pybytes()

    @staticmethod
    def _json_dataset(result: AgentTaskResult) -> bytes:
        payload = {
            "task_id": result.task_id,
            "unit_of_analysis": result.modeling_dataset.unit_of_analysis,
            "columns": [column.model_dump(mode="json") for column in result.modeling_dataset.columns],
            "rows": result.modeling_dataset.rows,
            "row_count": result.modeling_dataset.row_count,
            "source_datasets": [
                {
                    "name": item.name,
                    "dataset_role": item.dataset_role,
                    "study_key": item.study_key,
                    "row_count": item.row_count,
                    "columns": [column.model_dump(mode="json") for column in item.columns],
                    "rows": item.rows,
                }
                for item in result.source_datasets
            ],
        }
        return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")

    @staticmethod
    def _metadata(result: AgentTaskResult) -> bytes:
        payload = {
            "task_id": result.task_id,
            "question_parse": None if result.parsed_question is None else result.parsed_question.model_dump(mode="json"),
            "research_spec": result.research_spec.model_dump(mode="json"),
            "study_design": None if result.study_design is None else result.study_design.model_dump(mode="json"),
            "dictionary": [column.model_dump(mode="json") for column in result.modeling_dataset.columns],
            "sources": [item.model_dump(mode="json") for item in result.source_items],
        }
        return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")

    @staticmethod
    def _quality_report(result: AgentTaskResult) -> bytes:
        payload = {
            "task_id": result.task_id,
            "quality_gate": None if result.quality_gate_report is None else result.quality_gate_report.model_dump(mode="json"),
            "readiness": result.readiness.model_dump(mode="json"),
            "entity_matching": None if result.data_alignment is None else {
                "status": result.data_alignment.entity_match_status,
                "note": result.data_alignment.entity_match_note,
                "unresolved_identity_row_count": result.data_alignment.unresolved_identity_row_count,
                "cross_source_join_status": result.data_alignment.cross_source_join_status,
            },
            "cohort": None if result.cohort_construction is None else {
                "source_row_count": result.cohort_construction.source_row_count,
                "final_row_count": result.cohort_construction.final_row_count,
                "patient_count": result.cohort_construction.patient_count,
                "sample_count": result.cohort_construction.sample_count,
                "variable_coverage_rate": result.cohort_construction.variable_coverage_rate,
                "patient_linkage_f1": result.cohort_construction.patient_linkage_f1,
                "quality_gate": result.cohort_construction.quality_gate,
            },
        }
        return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")

    def _xlsx(self, result: AgentTaskResult) -> bytes:
        workbook = Workbook()
        data_sheet = workbook.active
        data_sheet.title = "科研数据集"
        columns = result.modeling_dataset.columns
        headers = [column.name for column in columns]
        data_sheet.append(headers)
        for row in result.modeling_dataset.rows:
            data_sheet.append([self._scalar(row.get(name)) for name in headers])
        self._style_table(data_sheet, len(headers))
        data_sheet.freeze_panes = "A2"
        data_sheet.auto_filter.ref = data_sheet.dimensions

        for index, companion in enumerate(result.source_datasets, start=1):
            title = AgentDatasetExportService._sheet_title(f"来源{index}_{companion.study_key or companion.name}")
            sheet = workbook.create_sheet(title)
            headers = [column.name for column in companion.columns]
            sheet.append(headers)
            for row in companion.rows:
                sheet.append([self._scalar(row.get(name)) for name in headers])
            if headers:
                self._style_table(sheet, len(headers))
                sheet.freeze_panes = "A2"
                sheet.auto_filter.ref = sheet.dimensions

        dictionary = workbook.create_sheet("字段字典")
        dictionary.append(["字段名", "中文标注", "数据类型", "科研用途", "来源字段", "说明"])
        for column in columns:
            dictionary.append(
                [
                    column.name,
                    column.label_zh,
                    column.data_type,
                    column.role,
                    column.source_field,
                    column.description,
                ]
            )
        self._style_table(dictionary, 6)
        dictionary.freeze_panes = "A2"

        report = workbook.create_sheet("可科研性报告")
        report.append(["项目", "内容"])
        entries: list[tuple[str, Any]] = [
            ("任务编号", result.task_id),
            ("科研问题", result.research_spec.research_goal),
            ("智能体模式", result.agent_mode),
            ("千问模型", result.model_name),
            ("分析单位", result.modeling_dataset.unit_of_analysis),
            ("数据行数", result.readiness.row_count),
            ("研究变量数", result.readiness.feature_count),
            ("研究结局字段", result.readiness.target_column or "未识别"),
            ("研究结局是否匹配", "是" if result.readiness.target_match else "否"),
            (
                "研究结局匹配率",
                "未计算" if result.readiness.target_match_rate is None else f"{result.readiness.target_match_rate:.1%}",
            ),
            (
                "全表字段完整率",
                "未计算" if result.readiness.field_completeness_rate is None else f"{result.readiness.field_completeness_rate:.1%}",
            ),
            (
                "主表基因变量覆盖率",
                "未指定" if result.readiness.requested_variable_coverage_rate is None else f"{result.readiness.requested_variable_coverage_rate:.1%}",
            ),
            ("自动清洗值数", result.readiness.cleaned_value_count),
            ("排除孤立分子记录数", result.readiness.excluded_orphan_record_count),
            ("是否支持科研分析", "是" if result.readiness.analysis_ready else "否"),
            ("分析分组建议", result.readiness.split_strategy),
            ("千问数据总结", result.summary_zh),
        ]
        for label, value in entries:
            report.append([label, value])
        for warning in result.readiness.warnings:
            report.append(["风险提示", warning])
        for action in result.readiness.cleaning_actions:
            report.append(["清洗动作", action])
        for recommendation in result.readiness.recommendations:
            report.append(["建议", recommendation])
        self._style_table(report, 2)
        report.column_dimensions["A"].width = 22
        report.column_dimensions["B"].width = 100
        report.column_dimensions["B"].alignment = Alignment(wrap_text=True, vertical="top")

        sources = workbook.create_sheet("数据来源")
        sources.append(["来源编号", "数据库", "数据编号（Accession）", "官方地址", "状态", "校验值"])
        for item in result.source_items:
            sources.append(
                [
                    item.source_id,
                    item.source_name,
                    item.accession,
                    item.url,
                    item.status,
                    item.checksum,
                ]
            )
        self._style_table(sources, 6)
        sources.freeze_panes = "A2"

        study_design = workbook.create_sheet("研究设计")
        study_design.append(["项目", "内容"])
        design = result.study_design
        if design is not None:
            for label, value in [
                ("研究类型", f"{design.research_type}｜{design.research_type_id}"),
                ("研究人群", design.population),
                ("核心暴露", design.exposure),
                ("研究结局", design.outcome),
                ("协变量", "、".join(design.covariates)),
                ("分析单位", design.analysis_unit),
                ("研究模型", design.model_expression),
                ("变量覆盖率", "未计算" if design.variable_coverage_rate is None else f"{design.variable_coverage_rate:.1%}"),
            ]:
                study_design.append([label, value])
            for rule in design.cohort_rules:
                study_design.append(["队列规则", rule])
            for variable in design.required_variables:
                study_design.append(
                    [
                        "变量",
                        f"{variable.variable_id}｜{variable.label}｜{variable.role}｜"
                        f"required={variable.required}｜available={variable.available}｜"
                        f"matched={','.join(variable.matched_fields) or '无'}｜{variable.note}",
                    ]
                )
            for source in design.data_source_recommendations:
                study_design.append(
                    [
                        "数据源建议",
                        f"{source.database}｜{source.availability}｜selected={source.selected}｜"
                        f"用途={source.purpose}｜域={','.join(source.data_domains)}｜{source.note}",
                    ]
                )
            for limitation in design.limitations:
                study_design.append(["限制", limitation])
        else:
            study_design.append(["状态", "当前任务没有返回研究设计报告。"])
        self._style_table(study_design, 2)
        study_design.column_dimensions["A"].width = 18
        study_design.column_dimensions["B"].width = 110
        study_design.column_dimensions["B"].alignment = Alignment(wrap_text=True, vertical="top")

        cohort_sheet = workbook.create_sheet("队列构建")
        cohort_sheet.append(["步骤", "规则类型", "筛选标准", "前置行数", "保留行数", "排除行数", "状态", "说明"])
        cohort = result.cohort_construction
        if cohort is not None:
            for step in cohort.filter_steps:
                cohort_sheet.append(
                    [
                        step.label,
                        step.rule_type,
                        step.criterion,
                        step.before_count,
                        step.after_count,
                        step.excluded_count,
                        step.status,
                        step.note,
                    ]
                )
            cohort_sheet.append([])
            cohort_sheet.append(["最终队列行数", cohort.final_row_count])
            cohort_sheet.append(["患者数", cohort.patient_count])
            cohort_sheet.append(["样本数", cohort.sample_count])
            cohort_sheet.append(["变量覆盖率", "未计算" if cohort.variable_coverage_rate is None else f"{cohort.variable_coverage_rate:.1%}"])
            cohort_sheet.append(["患者关联 F1", "未评测" if cohort.patient_linkage_f1 is None else f"{cohort.patient_linkage_f1:.3f}"])
            cohort_sheet.append(["Quality Gate", cohort.quality_gate])
            for note in cohort.notes:
                cohort_sheet.append(["说明", note])
        else:
            cohort_sheet.append(["状态", "当前任务没有返回队列构建报告。"])
        self._style_table(cohort_sheet, 8)
        cohort_sheet.freeze_panes = "A2"
        for column, width in {"A": 24, "B": 14, "C": 68, "D": 12, "E": 12, "F": 12, "G": 12, "H": 70}.items():
            cohort_sheet.column_dimensions[column].width = width
        for row in cohort_sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        collection_sheet = workbook.create_sheet("搜集智能体")
        collection_sheet.append(["项目", "内容"])
        collection = result.collection_agent
        if collection is not None:
            collection_sheet.append(["状态", collection.status])
            collection_sheet.append(["Quality Gate", collection.quality_gate])
            collection_sheet.append(["完成轮次", f"{collection.completed_rounds}/{collection.max_rounds}"])
            collection_sheet.append(["停止原因", collection.stop_reason or collection.note])
            collection_sheet.append(["诊断", collection.diagnosis or "未记录"])
            collection_sheet.append(["已试方法", "、".join(collection.strategies_tried) or "首轮规划"])
            for goal in collection.goals:
                collection_sheet.append(
                    [
                        "研究目标",
                        f"{goal.label}｜{'已达成' if goal.met else '未达成'}｜必需={goal.required}｜{goal.evidence}",
                    ]
                )
            collection_sheet.append(["说明", collection.note])
            for iteration in collection.iterations:
                collection_sheet.append(
                    [
                        f"第 {iteration.round_number} 轮",
                        f"{iteration.phase}｜状态={iteration.status}｜Gate={iteration.quality_gate}｜"
                        f"来源={','.join(iteration.source_names) or '无'}｜"
                        f"行={iteration.row_count}｜列={iteration.column_count}｜"
                        f"关键缺口={','.join(iteration.missing_critical_fields) or '无'}｜"
                        f"建议缺口={','.join(iteration.missing_recommended_fields) or '无'}｜"
                        f"{iteration.note}",
                    ]
                )
                for evidence in iteration.field_evidence:
                    collection_sheet.append(
                        [
                            "字段证据",
                            f"{evidence.source_name}｜来源={evidence.source_id or '未报告'}｜"
                            f"状态={evidence.status}｜字段={','.join(evidence.matched_fields) or '未匹配'}｜"
                            f"{evidence.note}",
                        ]
                    )
            for gap in collection.critical_gaps:
                collection_sheet.append(
                    [
                        "关键缺口",
                        f"{gap.variable_id}｜{gap.label}｜覆盖率="
                        f"{'未计算' if gap.coverage_rate is None else f'{gap.coverage_rate:.1%}'}｜"
                        f"建议来源={','.join(gap.suggested_sources)}｜{gap.reason}",
                    ]
                )
            for gap in collection.recommended_gaps:
                collection_sheet.append(
                    [
                        "建议缺口",
                        f"{gap.variable_id}｜{gap.label}｜覆盖率="
                        f"{'未计算' if gap.coverage_rate is None else f'{gap.coverage_rate:.1%}'}｜"
                        f"建议来源={','.join(gap.suggested_sources)}｜{gap.reason}",
                    ]
                )
            for action in collection.next_actions:
                collection_sheet.append(
                    [
                        "下一轮动作",
                        f"{action.source_name}｜{action.tool_name}｜优先级={action.priority}｜"
                        f"状态={action.status}｜参数={json.dumps(action.arguments, ensure_ascii=False)}｜"
                        f"{action.rationale}",
                    ]
                )
        else:
            collection_sheet.append(["状态", "当前任务没有返回搜集智能体报告。"])
        self._style_table(collection_sheet, 2)
        collection_sheet.column_dimensions["A"].width = 18
        collection_sheet.column_dimensions["B"].width = 120
        collection_sheet.column_dimensions["B"].alignment = Alignment(wrap_text=True, vertical="top")

        competition = workbook.create_sheet("比赛报告")
        competition.append(["项目", "内容"])
        report = result.competition_report
        if report is not None:
            competition_rows = [
                ("赛道", report.track),
                ("方向", report.direction),
                ("聚焦问题", report.problem_focus),
                ("总体摘要", report.summary),
            ]
            for label, value in competition_rows:
                competition.append([label, value])
            if report.unified_evaluation is not None:
                unified = report.unified_evaluation
                competition.append(["统一评价体系", f"{unified.version}｜{unified.status}"])
                competition.append(["统一评价体系声明", unified.no_fake_scores_notice])
                for layer in unified.layers:
                    competition.append(
                        [
                            "评价层级",
                            f"{layer.layer_id}｜{layer.label}｜{layer.status}｜"
                            f"{layer.purpose}｜输出：{', '.join(layer.primary_outputs)}｜"
                            f"证据要求：{layer.evidence_requirement}",
                        ]
                    )
                fitness = unified.task_adaptive_fitness
                competition.append(
                    [
                        "Task-Adaptive Fitness",
                        f"{fitness.evaluation_contract_id}｜冻结={fitness.frozen_before_run}｜"
                        f"状态={fitness.status}｜Fitness={fitness.fitness_score if fitness.fitness_score is not None else '未计算'}｜"
                        f"Gate={fitness.quality_gate}｜发布={fitness.publish_allowed}｜{fitness.note}",
                    ]
                )
                for dimension in fitness.dimensions:
                    competition.append(
                        [
                            "Fitness维度",
                            f"{dimension.name}｜{dimension.display_value}｜{dimension.status}｜{dimension.detail}",
                        ]
                    )
                for gap in fitness.gap_feedback:
                    competition.append(["Fitness缺口", gap])
                if unified.model_comparison:
                    for row in unified.model_comparison:
                        competition.append(
                            [
                                "模型对比",
                                f"{row.method_id}｜{row.method_label}｜model={row.base_model_id or 'N/A'}｜"
                                f"status={row.status}｜fitness={row.fitness_score if row.fitness_score is not None else '待实测'}｜"
                                f"sdti={row.sdti_status}｜gate={row.quality_gate}｜publish={row.publish_allowed}｜{row.note}",
                            ]
                        )
                else:
                    competition.append(
                        [
                            "模型对比",
                            "未运行横向评价；当前任务级科研适配度不作为模型分数。"
                            "请在独立模型评价页使用同一批问题完成至少两个模型/变体的真实测试。",
                        ]
                    )
                for table in unified.horizontal_comparisons:
                    competition.append(["横向对比", f"{table.table_id}｜{table.title}｜{table.status}｜{table.note}"])
                    for table_row in table.rows:
                        competition.append(
                            [
                                "横向对比行",
                                f"{table.table_id}｜"
                                + "｜".join(f"{key}={value}" for key, value in table_row.items()),
                            ]
                        )
                for row in unified.stratified_comparisons:
                    metrics_text = "；".join(f"{key}={value}" for key, value in row.metrics.items())
                    competition.append(
                        [
                            "分层对比",
                            f"{row.stratum_name}={row.stratum_value}｜n={row.n}｜"
                            f"gate={row.quality_gate}｜publish={row.publish_allowed}｜{metrics_text}｜{row.note}",
                        ]
                    )
                for next_run in unified.required_next_runs:
                    competition.append(["待补实验", next_run])
            for metric in report.metrics:
                competition.append(["指标", f"{metric.name}｜{metric.display_value}｜{metric.target}｜{metric.detail}"])
            for row in report.ablation_rows:
                score = "未计算" if row.diagnostic_score is None else f"{row.diagnostic_score:.1f}"
                delta = "" if row.delta_from_full is None else f"｜delta={row.delta_from_full:.1f}"
                competition.append(["消融", f"{row.variant}｜{score}{delta}｜{row.removed_component}｜{row.expected_effect}｜{row.observed_effect}｜{row.note}"])
            for row in report.variant_scores:
                score = "未计算" if row.diagnostic_score is None else f"{row.diagnostic_score:.1f}"
                competition.append(["对照变体", f"{row.label}｜{score}｜{row.status}｜{row.note}"])
            for layer in report.rag_layers:
                competition.append(["混合RAG", f"{layer.layer}｜{layer.implementation}｜{layer.why_it_matters}｜{layer.observable_effect}"])
            for node in report.rag_flow_nodes:
                competition.append(["RAG流程节点", f"{node.order}｜{node.layer}｜{node.label}｜{node.status}｜{node.detail}"])
            for edge in report.rag_flow_edges:
                competition.append(["RAG流程边", f"{edge.source} -> {edge.target}｜{edge.label}｜{edge.detail or ''}"])
            for match in report.rag_matches:
                signal_text = "；".join(f"{name}={value:.0%}" for name, value in match.signals.items())
                competition.append(
                    [
                        "RAG库匹配",
                        f"{match.database}｜{match.dataset_name}｜{match.display_score}｜{match.status}｜"
                        f"{'已选用' if match.selected else '候选'}｜{signal_text}｜{match.rationale}",
                    ]
                )
            competition.append(["知识图谱", f"节点 {report.knowledge_graph.node_count}｜边 {report.knowledge_graph.edge_count}｜关系 {', '.join(report.knowledge_graph.relation_types)}"])
            competition.append(["知识图谱", report.knowledge_graph.note])
            for node in report.graph_nodes:
                competition.append(["知识图谱节点", f"{node.node_id}｜{node.label}｜{node.node_type}｜{node.group}｜{node.status or ''}｜{node.detail or ''}"])
            for edge in report.graph_edges:
                competition.append(["知识图谱边", f"{edge.source} -> {edge.target}｜{edge.label}｜{edge.relation_type}｜strength={edge.strength:.2f}｜{edge.detail or ''}"])
            if report.scientific_usability is not None:
                analysis = report.scientific_usability
                competition.append(["科研适用性", f"{analysis.status}｜样本 {analysis.sample_size}｜结局 {analysis.target_column or '未识别'}｜特征 {analysis.feature_count}｜方法 {', '.join(analysis.methods)}"])
                competition.append(["科研适用性", analysis.interpretation])
                for finding in analysis.findings:
                    competition.append(["科研适用性发现", f"{finding.variable} -> {finding.outcome}｜{finding.method}｜n={finding.n}｜{finding.display_score}｜{finding.status}｜{finding.interpretation}"])
                for caveat in analysis.caveats:
                    competition.append(["科研适用性注意", caveat])
            for item in report.improvement_highlights:
                competition.append(["改进", item])
            for item in report.limitations:
                competition.append(["局限", item])
            for item in report.submission_checklist:
                competition.append(["提交核验", f"{item.label}｜{item.status}｜{item.detail}"])
            for item in report.deliverables:
                competition.append(["交付物", item])
        self._style_table(competition, 2)
        competition.column_dimensions["A"].width = 18
        competition.column_dimensions["B"].width = 110
        competition.column_dimensions["B"].alignment = Alignment(wrap_text=True, vertical="top")

        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    @staticmethod
    def _sheet_title(name: str) -> str:
        cleaned = "".join(ch if ch not in r"[]:*?/\ " else "_" for ch in str(name or "来源表"))
        return cleaned[:31] or "来源表"

    @staticmethod
    def _style_table(sheet: Any, column_count: int) -> None:
        fill = PatternFill("solid", fgColor="0F766E")
        for cell in sheet[1]:
            cell.fill = fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_index in range(1, column_count + 1):
            values = [str(sheet.cell(row=row, column=column_index).value or "") for row in range(1, min(sheet.max_row, 100) + 1)]
            width = min(max(max((len(value) for value in values), default=8) + 2, 12), 42)
            sheet.column_dimensions[get_column_letter(column_index)].width = width
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=False)

    @staticmethod
    def _scalar(value: Any) -> Any:
        if isinstance(value, (dict, list, tuple, set)):
            return json.dumps(value, ensure_ascii=False, sort_keys=True)
        return value
