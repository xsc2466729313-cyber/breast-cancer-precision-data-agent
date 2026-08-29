from __future__ import annotations

import json
import gzip
import time
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

import pytest

import httpx
import pyarrow as pa
import pyarrow.parquet as pq
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend.app.agent import (
    AgentDatasetExportService,
    AgentExportFormat,
    AgentTaskRequest,
    QwenClient,
    QwenSettings,
    ResearchAgentService,
)
from backend.app.agent.collection_agent import CollectionAgent
from backend.app.agent.models import CollectionGap, AnalysisReadinessReport, ModelingDataset
from backend.app.agent.models import DatasetColumn
from backend.app.agent.dataset_builder import ResearchDatasetBuilder
from backend.app.agent.study_design import StudyDesignBuilder
from backend.app.main import app
from backend.app.models import ResearchSpec, SourceItem
from backend.app.sources.cbioportal import CBioPortalAdapter
from backend.app.sources.discovery import DiscoveryAdapter
from backend.app.sources.geo.models import (
    GEOAdapterResult,
    GEOCacheStatus,
    GEOResourceRecord,
    GEOResourceType,
)
from backend.tests.test_cbioportal_adapter import standard_handler


QUESTION = "研究 HER2 阳性乳腺癌中 PIK3CA 突变与治疗响应的关系"


def qwen_handler(request: httpx.Request) -> httpx.Response:
    payload = json.loads(request.content)
    has_tool_result = any(message.get("role") == "tool" for message in payload["messages"])
    if payload.get("response_format") and not has_tool_result:
        content = json.dumps(
            {
                "research_goal": QUESTION,
                "disease": "Breast Cancer",
                "subtype": "HER2-positive",
                "genes": ["ERBB2", "PIK3CA", "TP53"],
                "variants": [],
                "drugs": [],
                "outcomes": ["treatment_response"],
                "required_data_types": ["clinical", "mutation", "treatment_response"],
                "target_fields": ["patient_id", "sample_id", "response"],
            },
            ensure_ascii=False,
        )
        message = {"role": "assistant", "content": content}
    elif payload.get("tools"):
        message = {
            "role": "assistant",
            "content": None,
            "tool_calls": [
                {
                    "id": "call-cbio",
                    "type": "function",
                    "function": {
                        "name": "search_cbioportal",
                        "arguments": json.dumps(
                            {
                                "study_id": "brca_metabric",
                                "gene_symbols": ["ERBB2", "PIK3CA", "TP53"],
                                "max_records": 100,
                            }
                        ),
                    },
                }
            ],
        }
    else:
        assert has_tool_result
        message = {
            "role": "assistant",
            "content": json.dumps({"summary": "已获取 METABRIC 患者级临床、突变和拷贝数记录；当前样例记录数较少，需扩大队列并补充治疗响应标签。"}, ensure_ascii=False),
        }
    return httpx.Response(
        200,
        json={"model": "qwen-plus", "choices": [{"message": message}]},
        request=request,
    )


def discovery_handler(request: httpx.Request) -> httpx.Response:
    if request.url.path.endswith("/esearch.fcgi"):
        return httpx.Response(
            200,
            json={"esearchresult": {"count": "0", "idlist": []}},
            request=request,
        )
    if request.url.path.endswith("/esummary.fcgi"):
        return httpx.Response(200, json={"result": {"uids": []}}, request=request)
    if request.url.host == "www.ebi.ac.uk":
        return httpx.Response(
            200,
            json={"hitCount": 0, "resultList": {"result": []}},
            request=request,
        )
    raise AssertionError(f"Unexpected discovery request: {request.url}")


def build_agent(tmp_path: Path) -> ResearchAgentService:
    qwen_http = httpx.Client(transport=httpx.MockTransport(qwen_handler))
    qwen = QwenClient(
        settings=QwenSettings(
            api_key="test-key",
            base_url="https://ws-test.cn-beijing.maas.aliyuncs.com/compatible-mode/v1",
            model="qwen-plus",
            workspace_id="ws-test",
        ),
        client=qwen_http,
    )
    cbio_http = httpx.Client(transport=httpx.MockTransport(standard_handler))
    cbio = CBioPortalAdapter(cache_dir=tmp_path / "cbio", client=cbio_http)
    discovery = DiscoveryAdapter(client=httpx.Client(transport=httpx.MockTransport(discovery_handler)))
    return ResearchAgentService(qwen_client=qwen, cbioportal_adapter=cbio, discovery_adapter=discovery)


def test_qwen_agent_executes_function_call_and_builds_research_table(tmp_path: Path) -> None:
    result = build_agent(tmp_path).run(
        AgentTaskRequest(
            question=QUESTION,
            use_qwen=True,
            allow_deterministic_fallback=False,
            data_mode="live",
            max_sources=1,
            max_records=100,
            iterative_collection=False,
        )
    )

    assert result.used_qwen is True
    assert result.used_model is True
    assert result.agent_mode == "千问科研数据智能体"
    assert result.model_provider == "千问"
    assert result.tool_calls[0].tool_name == "search_cbioportal"
    assert result.tool_calls[0].status == "完成"
    assert result.modeling_dataset.row_count == 1
    names = {column.name for column in result.modeling_dataset.columns}
    assert {"patient_id", "sample_id", "her2_status", "erbb2_cna"}.issubset(names)
    assert "pik3ca_mutation" not in names
    assert result.modeling_dataset.target_column is None
    assert result.readiness.target_match is False
    assert result.readiness.excluded_orphan_record_count == 4
    labels = {column.name: column.label_zh for column in result.modeling_dataset.columns}
    assert labels["her2_status"] == "HER2 状态"
    assert result.readiness.status == "研究结局不匹配"
    assert "METABRIC" in result.summary_zh
    assert result.source_items
    assert result.collection_agent is not None
    assert result.collection_agent.completed_rounds >= 1
    assert any(
        gap.variable_id in {"sample_type", "sample_timepoint", "treatment", "outcome"}
        for gap in result.collection_agent.critical_gaps
    )


def test_collection_agent_deduplicates_search_requests_by_arguments_not_tool_name() -> None:
    agent = CollectionAgent()
    spec = ResearchSpec(
        task_id="collection-test",
        research_goal="研究乳腺癌治疗响应",
        disease="Breast Cancer",
        genes=["PIK3CA"],
        outcomes=["treatment_response"],
        required_data_types=["clinical", "mutation", "treatment_response"],
        target_fields=["patient_id", "sample_id"],
    )
    gaps = [
        CollectionGap(
            variable_id="sample_type",
            label="样本类型",
            role="样本信息",
            required=True,
            coverage_rate=0.0,
            reason="缺失",
        )
    ]

    first = agent.propose_actions(
        spec=spec,
        gaps=gaps,
        attempted_calls=set(),
        max_records=100,
    )
    attempted = {
        agent.call_key({"name": action.tool_name, "arguments": action.arguments})
        for action in first
    }
    second = agent.propose_actions(
        spec=spec,
        gaps=gaps,
        attempted_calls=attempted,
        max_records=100,
    )

    assert first
    assert all(
        agent.call_key({"name": action.tool_name, "arguments": action.arguments})
        not in attempted
        for action in second
    )
    assert any(action.tool_name == "search_cbioportal" for action in first)


def test_her2_positive_response_gap_searches_gse76360_first() -> None:
    agent = CollectionAgent()
    spec = ResearchSpec(
        task_id="geo-order-test",
        research_goal="研究 HER2 阳性乳腺癌治疗响应",
        disease="Breast Cancer",
        subtype="HER2-positive",
        genes=["PIK3CA"],
        outcomes=["treatment_response"],
        required_data_types=["clinical", "mutation", "treatment_response"],
    )
    gaps = [
        CollectionGap(
            variable_id="outcome",
            label="研究结局",
            role="结局",
            required=True,
            coverage_rate=0.0,
            reason="缺失治疗响应",
        )
    ]

    actions = agent.propose_actions(spec=spec, gaps=gaps, attempted_calls=set(), max_records=100)
    geo_actions = [action for action in actions if action.tool_name == "search_geo"]

    assert geo_actions
    assert geo_actions[0].arguments["accession"] == "GSE76360"


def test_response_question_prefers_outcome_matched_cohort_over_larger_molecular_table() -> None:
    spec = ResearchSpec(
        task_id="select-test",
        research_goal="研究 HER2 阳性乳腺癌中 PIK3CA 突变是否影响治疗响应",
        disease="Breast Cancer",
        subtype="HER2-positive",
        genes=["PIK3CA"],
        outcomes=["treatment_response"],
        required_data_types=["clinical", "mutation", "treatment_response"],
    )
    molecular, molecular_ready = ResearchDatasetBuilder().empty()
    molecular = molecular.model_copy(update={"name": "METABRIC", "row_count": 848})
    molecular_ready = molecular_ready.model_copy(
        update={
            "target_match": False,
            "requested_variable_coverage_rate": 0.7,
            "field_completeness_rate": 0.689,
        }
    )
    response, response_ready = ResearchDatasetBuilder().empty()
    response = response.model_copy(update={"name": "GSE76360", "row_count": 50})
    response_ready = response_ready.model_copy(
        update={
            "target_match": True,
            "requested_variable_coverage_rate": 0.0,
            "field_completeness_rate": 0.9,
        }
    )

    chosen, _ready = max(
        [(molecular, molecular_ready), (response, response_ready)],
        key=lambda item: ResearchAgentService._dataset_selection_score(item, spec),
    )

    assert chosen.name == "GSE76360"


def test_selection_prefers_same_patient_gene_and_response_pack() -> None:
    spec = ResearchSpec(
        task_id="dual-select",
        research_goal="研究 HER2 阳性乳腺癌中 PIK3CA 突变是否影响治疗响应",
        disease="Breast Cancer",
        subtype="HER2-positive",
        genes=["PIK3CA"],
        outcomes=["treatment_response"],
        required_data_types=["clinical", "mutation", "treatment_response"],
    )
    response_only, response_ready = ResearchDatasetBuilder().empty()
    response_only = response_only.model_copy(update={"name": "GSE76360", "row_count": 50})
    response_ready = response_ready.model_copy(
        update={"target_match": True, "requested_variable_coverage_rate": 0.0, "field_completeness_rate": 0.9}
    )
    dual, dual_ready = ResearchDatasetBuilder().empty()
    dual = dual.model_copy(update={"name": "breast_alpelisib_2020", "row_count": 40})
    dual_ready = dual_ready.model_copy(
        update={"target_match": True, "requested_variable_coverage_rate": 1.0, "field_completeness_rate": 0.85}
    )

    chosen, _ready = max(
        [(response_only, response_ready), (dual, dual_ready)],
        key=lambda item: ResearchAgentService._dataset_selection_score(item, spec),
    )

    assert chosen.name == "breast_alpelisib_2020"


def test_iterative_collection_switches_to_geo_response_cohort(tmp_path: Path) -> None:
    result = build_agent(tmp_path).run(
        AgentTaskRequest(
            question=QUESTION,
            use_qwen=True,
            allow_deterministic_fallback=False,
            data_mode="live",
            max_sources=1,
            max_records=100,
            iterative_collection=True,
            max_collection_rounds=3,
        )
    )

    assert result.readiness.target_match is True
    assert result.modeling_dataset.target_column == "treatment_response"
    assert result.modeling_dataset.row_count >= 30
    assert "曲妥珠" in result.modeling_dataset.name or "GSE76360" in result.modeling_dataset.name
    assert any("切换到含治疗响应" in warning for warning in result.readiness.warnings)
    assert result.cohort_construction is not None
    assert result.cohort_construction.final_row_count > 0
    assert result.collection_agent is not None
    assert result.collection_agent.quality_gate in {"PARTIAL", "REVIEW", "PASS"}
    assert result.collection_agent.completed_rounds >= 2
    assert any(item.diagnosis == "outcome_mismatch" for item in result.collection_agent.iterations)
    assert any(
        "GSE76360" in (item.note or "") or "gse76360" in " ".join(item.strategy_ids)
        for item in result.collection_agent.iterations
    )
    assert not any("跨库" in (warning or "") and "已把" in (warning or "") for warning in result.readiness.warnings)


def test_autonomous_follow_up_converts_catalog_hits_into_geo_fetch() -> None:
    from types import SimpleNamespace

    service = ResearchAgentService()
    spec = ResearchSpec(
        task_id="harvest-follow-up",
        research_goal="研究 HER2 阳性乳腺癌中 PIK3CA 突变是否影响治疗响应",
        disease="Breast Cancer",
        subtype="HER2-positive",
        genes=["PIK3CA"],
        outcomes=["treatment_response"],
        required_data_types=["clinical", "mutation", "treatment_response"],
    )
    catalog = SimpleNamespace(
        records=[
            SimpleNamespace(
                accession="GSE50948",
                title="HER2 positive breast cancer PIK3CA trastuzumab response",
                summary="pCR",
            )
        ]
    )
    decision = SimpleNamespace(
        actions=[],
        diagnosis="missing_same_cohort_exposure",
        quality_gate="REVIEW",
        goals=[],
    )
    calls = service._autonomous_follow_up_calls(
        spec=spec,
        request=AgentTaskRequest(question=QUESTION, use_qwen=False),
        decision=decision,
        raw_results=[("search_geo_catalog", catalog)],
        critical=[],
        dataset=SimpleNamespace(row_count=50),
        readiness=SimpleNamespace(target_match=True),
        attempted_calls=set(),
        qwen_client=None,
        round_number=2,
        max_rounds=8,
    )

    assert calls
    accessions = [str(call["arguments"].get("accession") or "") for call in calls if call["name"] == "search_geo"]
    assert "GSE50948" in accessions


def test_hr_positive_her2_negative_pi3k_question_skips_her2_geo_response_cohort() -> None:
    question = "PIK3CA 突变的 HR+/HER2- 乳腺癌患者，使用 PI3K 抑制剂后的响应是否优于野生型？"
    request = AgentTaskRequest(
        question=question,
        use_qwen=False,
        data_mode="live",
        max_sources=5,
        max_records=500,
    )
    service = ResearchAgentService()
    spec = service._enrich_research_spec(
        service._deterministic_spec(question, "task-hr-her2-pi3k"),
        question,
    )
    calls = service._deterministic_tool_calls(spec, request)
    guarded = service._guard_tool_arguments(
        [{"id": "qwen-wrong-geo", "name": "search_geo", "arguments": {"accession": "GSE76360"}}, *calls],
        spec,
        request,
    )

    assert spec.subtype == "HR-positive/HER2-negative"
    assert spec.genes == ["PIK3CA"]
    assert spec.drugs == ["Alpelisib"]
    assert "search_geo" not in {call["name"] for call in guarded}
    cbio_call = next(call for call in guarded if call["name"] == "search_cbioportal")
    assert cbio_call["arguments"]["gene_symbols"] == ["PIK3CA"]
    civic_call = next(call for call in guarded if call["name"] == "search_civic")
    assert civic_call["arguments"]["therapy_name"] == "Alpelisib"


def test_tnbc_pcr_question_prioritizes_response_geo_and_literature() -> None:
    question = "研究三阴性乳腺癌中 BRCA1/BRCA2 突变与新辅助化疗病理完全缓解（pCR）的关系，并整理患者级科研数据集"
    request = AgentTaskRequest(
        question=question,
        use_qwen=False,
        data_mode="plan_only",
        max_sources=5,
        max_records=200,
    )
    service = ResearchAgentService()
    spec = service._enrich_research_spec(service._deterministic_spec(question, "task-tnbc-pcr"), question)
    calls = service._deterministic_tool_calls(spec, request)
    names = [call["name"] for call in calls]

    assert spec.subtype == "Triple-negative"
    assert spec.genes == ["BRCA1", "BRCA2"]
    assert "treatment_response" in spec.outcomes or "pCR" in " ".join(spec.outcomes)
    assert names[0] == "search_geo"
    assert calls[0]["arguments"]["accession"] == "GSE25066"
    assert "search_geo_catalog" in names
    assert "search_europe_pmc" in names


def test_agent_excel_export_contains_chinese_dictionary_and_readiness(tmp_path: Path) -> None:
    result = build_agent(tmp_path).run(
        AgentTaskRequest(
            question=QUESTION,
            use_qwen=True,
            allow_deterministic_fallback=False,
            data_mode="live",
            max_sources=1,
            max_records=100,
            iterative_collection=False,
        )
    )
    exported = AgentDatasetExportService().export(result, AgentExportFormat.XLSX)
    workbook = load_workbook(BytesIO(exported.content), read_only=True)

    assert workbook.sheetnames == [
        "科研数据集",
        "字段字典",
        "可科研性报告",
        "数据来源",
        "研究设计",
        "队列构建",
        "搜集智能体",
        "比赛报告",
    ]
    assert workbook["科研数据集"]["A1"].value == "study_id"
    assert workbook["字段字典"]["B1"].value == "中文标注"
    assert workbook["字段字典"]["D1"].value == "科研用途"
    assert workbook["可科研性报告"]["A2"].value == "任务编号"
    assert workbook["研究设计"]["A1"].value == "项目"
    assert workbook["队列构建"]["A1"].value == "步骤"
    assert workbook["搜集智能体"]["A1"].value == "项目"
    competition_values = [cell.value for row in workbook["比赛报告"].iter_rows(values_only=False) for cell in row if cell.value]
    assert "科研适用性" in competition_values
    assert "统一评价体系" in competition_values
    assert "模型对比" in competition_values
    assert "横向对比" in competition_values
    assert "分层对比" in competition_values
    assert "RAG流程节点" in competition_values
    assert "RAG库匹配" in competition_values
    assert "知识图谱节点" in competition_values


def test_agent_parquet_export_normalizes_mixed_upstream_types(tmp_path: Path) -> None:
    result = build_agent(tmp_path).run(
        AgentTaskRequest(
            question=QUESTION,
            use_qwen=True,
            allow_deterministic_fallback=False,
            data_mode="live",
            max_sources=1,
            max_records=100,
            iterative_collection=False,
        )
    )
    if not any(column.name == "stage" for column in result.modeling_dataset.columns):
        result.modeling_dataset.columns.append(
            DatasetColumn(
                name="stage",
                label_zh="肿瘤分期",
                data_type="string",
                role="研究变量",
                source_field="STAGE",
                description="混合上游类型测试",
            )
        )
    result.modeling_dataset.rows.append(dict(result.modeling_dataset.rows[0]))
    result.modeling_dataset.rows[0]["stage"] = 2
    result.modeling_dataset.rows[1]["stage"] = "Stage III"
    result.modeling_dataset.row_count = len(result.modeling_dataset.rows)
    exported = AgentDatasetExportService().export(result, AgentExportFormat.PARQUET)
    table = pq.read_table(pa.BufferReader(exported.content))

    assert table.num_rows == result.modeling_dataset.row_count
    assert table["stage"].type == pa.string()
    assert table["stage"][0].as_py() == "2"


def test_agent_api_plan_only_is_not_mock() -> None:
    client = TestClient(app)
    response = client.post(
        "/api/agent/tasks",
        json={
            "question": QUESTION,
            "use_qwen": False,
            "data_mode": "plan_only",
            "max_sources": 2,
            "max_records": 100,
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["agent_mode"] == "确定性科研规划"
    assert payload["used_qwen"] is False
    assert payload["plan"][1]["label"] == "生成研究设计与数据规划"
    assert payload["parsed_question"]["disease"]
    assert payload["parsed_question"]["required_variables"]
    assert payload["quality_gate_report"]["overall"] in {"PASS", "REVIEW", "REJECT"}
    assert len(payload["quality_gate_report"]["layers"]) == 4
    assert payload["quality_gate_report"]["cohort_f1"] is None
    assert "Mock" not in payload["notice"]


def test_geo_series_matrix_builds_baseline_response_cohort(tmp_path: Path) -> None:
    matrix = tmp_path / "GSE76360_series_matrix.txt.gz"
    lines = [
        '!Sample_title\t"119_B"\t"119_P"\t"120_B"\t"120_P"',
        '!Sample_geo_accession\t"GSM1"\t"GSM2"\t"GSM3"\t"GSM4"',
        '!Sample_characteristics_ch1\t"subject id: 119"\t"subject id: 119"\t"subject id: 120"\t"subject id: 120"',
        '!Sample_characteristics_ch1\t"patient status: HER2+ Breast Cancer"\t"patient status: HER2+ Breast Cancer"\t"patient status: HER2+ Breast Cancer"\t"patient status: HER2+ Breast Cancer"',
        '!Sample_characteristics_ch1\t"timepoint: baseline"\t"timepoint: post"\t"timepoint: baseline"\t"timepoint: post"',
        '!Sample_characteristics_ch1\t"response at surgery: pCR"\t"response at surgery: pCR"\t"response at surgery: NOR"\t"response at surgery: NOR"',
        '!Sample_characteristics_ch1\t"er status: Pos"\t"er status: Pos"\t"er status: Neg"\t"er status: Neg"',
        '!Sample_characteristics_ch1\t"pr status: Neg"\t"pr status: Neg"\t"pr status: Neg"\t"pr status: Neg"',
        '!series_matrix_table_begin',
    ]
    with gzip.open(matrix, "wt", encoding="utf-8") as handle:
        handle.write("\n".join(lines))
    source = SourceItem(
        source_id="geo:GSE76360:test",
        task_id="task-geo",
        source_name="NCBI GEO",
        source_type="database",
        accession="GSE76360",
        url="https://ftp.ncbi.nlm.nih.gov/test",
        file_type="series_matrix",
        local_path=str(matrix),
        checksum="sha256:test",
        status="downloaded",
    )
    geo = GEOAdapterResult(
        task_id="task-geo",
        accession="GSE76360",
        portal_url="https://www.ncbi.nlm.nih.gov/geo/query/acc.cgi?acc=GSE76360",
        availability=[],
        resources=[
            GEOResourceRecord(
                accession="GSE76360",
                resource_type=GEOResourceType.SERIES_MATRIX,
                file_name=matrix.name,
                download_url=source.url,
                status="downloaded",
                file_size=matrix.stat().st_size,
                source_item=source,
            )
        ],
        source_items=[source],
        cache_hit=GEOCacheStatus(accession_directory=False, resource_directories={}),
        queried_at=datetime.now(timezone.utc),
        notice="test",
    )
    spec = ResearchSpec(
        task_id="task-geo",
        research_goal=QUESTION,
        disease="Breast Cancer",
        subtype="HER2-positive",
        genes=["PIK3CA"],
        outcomes=["treatment_response"],
        required_data_types=["clinical", "treatment_response"],
    )

    built = ResearchDatasetBuilder().build_from_geo(geo, spec)

    assert built is not None
    dataset, readiness = built
    assert dataset.name == "HER2 阳性乳腺癌术前曲妥珠单抗响应队列科研数据集"
    assert dataset.row_count == 2
    assert dataset.patient_count == 2
    assert dataset.target_column == "treatment_response"
    assert dataset.class_distribution == {"病理完全缓解（pCR）": 1, "未达客观缓解": 1}
    assert {row["timepoint"] for row in dataset.rows} == {"基线"}
    assert all(row["raw_characteristics"] for row in dataset.rows)
    assert all(row["subtype"] == "HER2-positive" for row in dataset.rows)
    assert all(row["her2_status"] == "阳性" for row in dataset.rows)
    assert all(row["treatment"] == "曲妥珠单抗新辅助治疗" for row in dataset.rows)
    assert all(row["sample_type"] == "原发肿瘤" for row in dataset.rows)
    assert all(row["sample_source"] == "乳腺肿瘤穿刺活检" for row in dataset.rows)
    assert all(row["sample_timepoint"] == "基线" for row in dataset.rows)
    assert all(row["disease"] == "乳腺癌" for row in dataset.rows)
    assert readiness.target_match is True
    assert readiness.target_missing_rate == 0
    assert readiness.requested_variable_coverage_rate == 0
    assert any("治疗后配对样本" in action for action in readiness.cleaning_actions)


def test_gse25066_maps_pcr_rd_and_her2_ihc_without_promoting_two_plus(tmp_path: Path) -> None:
    matrix = tmp_path / "GSE25066_series_matrix.txt.gz"
    lines = [
        '!Sample_title\t"A"\t"B"\t"C"',
        '!Sample_geo_accession\t"GSM1"\t"GSM2"\t"GSM3"',
        '!Sample_characteristics_ch1\t"subject id: 1"\t"subject id: 2"\t"subject id: 3"',
        '!Sample_characteristics_ch1\t"pathologic_response_pcr_rd: pCR"\t"pathologic_response_pcr_rd: RD"\t"pathologic_response_pcr_rd: pCR"',
        '!Sample_characteristics_ch1\t"her2_status_ihc: 3+"\t"her2_status_ihc: 2+"\t"her2_status_ihc: Neg"',
        '!series_matrix_table_begin',
    ]
    with gzip.open(matrix, "wt", encoding="utf-8") as handle:
        handle.write("\n".join(lines))
    source = SourceItem(
        source_id="geo:GSE25066:test",
        task_id="task-gse25066",
        source_name="NCBI GEO",
        source_type="database",
        accession="GSE25066",
        url="https://ftp.ncbi.nlm.nih.gov/test",
        file_type="series_matrix",
        local_path=str(matrix),
        checksum="sha256:test",
        status="downloaded",
    )
    geo = GEOAdapterResult(
        task_id="task-gse25066",
        accession="GSE25066",
        portal_url="https://www.ncbi.nlm.nih.gov/geo/query/acc.cgi?acc=GSE25066",
        availability=[],
        resources=[
            GEOResourceRecord(
                accession="GSE25066",
                resource_type=GEOResourceType.SERIES_MATRIX,
                file_name=matrix.name,
                download_url=source.url,
                status="downloaded",
                file_size=matrix.stat().st_size,
                source_item=source,
            )
        ],
        source_items=[source],
        cache_hit=GEOCacheStatus(accession_directory=False, resource_directories={}),
        queried_at=datetime.now(timezone.utc),
        notice="test",
    )
    spec = ResearchSpec(
        task_id="task-gse25066",
        research_goal="研究乳腺癌新辅助治疗病理完全缓解（pCR）与 HER2 状态",
        disease="Breast Cancer",
        genes=[],
        outcomes=["pCR"],
        required_data_types=["clinical", "treatment_response"],
    )
    built = ResearchDatasetBuilder().build_from_geo(geo, spec)
    assert built is not None
    dataset, readiness = built
    assert dataset.target_column in {"pcr", "treatment_response"}
    assert readiness.target_match is True
    her2 = [row.get("her2_status") for row in dataset.rows]
    assert any(value == "2+" for value in her2)
    assert ResearchDatasetBuilder._receptor_polarity("2+") == "equivocal"
    assert all(row.get("pcr") or row.get("treatment_response") for row in dataset.rows)


def test_focus_accessions_seed_geo_calls_when_max_sources_is_two() -> None:
    service = ResearchAgentService()
    request = AgentTaskRequest(
        question="研究乳腺癌新辅助治疗病理完全缓解（pCR）需要患者级结局",
        use_qwen=False,
        data_mode="plan_only",
        max_sources=2,
        max_records=100,
        focus_accessions=["GSE25066", "GSE76360"],
        iterative_collection=False,
    )
    spec = service._deterministic_spec(request.question, "focus-seed-test")
    spec = service._enrich_research_spec(spec, request.question)
    calls = service._priority_seed_calls(spec, request)
    guarded = service._guard_tool_arguments(calls, spec, request)
    accessions = [str((call.get("arguments") or {}).get("accession") or "") for call in guarded]
    assert "GSE25066" in accessions
    assert "GSE76360" in accessions or accessions[0] == "GSE25066"


def test_plan_only_exports_metadata_and_quality_report_without_rows() -> None:
    result = ResearchAgentService().run(
        AgentTaskRequest(
            question=QUESTION,
            use_qwen=False,
            data_mode="plan_only",
            max_sources=2,
            max_records=100,
        )
    )
    exporter = AgentDatasetExportService()
    metadata = json.loads(exporter.export(result, AgentExportFormat.METADATA).content)
    report = json.loads(exporter.export(result, AgentExportFormat.QUALITY_REPORT).content)

    assert metadata["question_parse"]["disease"]
    assert metadata["research_spec"]["disease"]
    assert report["quality_gate"]["overall"] in {"PASS", "REVIEW", "REJECT"}
    assert len(report["quality_gate"]["layers"]) == 4
    assert report["entity_matching"]["status"] in {"MATCH", "REVIEW", "UNMATCH"}
    with pytest.raises(ValueError, match="没有可导出的科研数据行"):
        exporter.export(result, AgentExportFormat.JSON)


def test_research_task_api_polls_status_and_returns_spec_and_report() -> None:
    client = TestClient(app)
    created = client.post(
        "/api/research/task",
        json={
            "question": QUESTION,
            "use_qwen": False,
            "data_mode": "plan_only",
            "max_sources": 2,
            "max_records": 100,
        },
    )

    assert created.status_code == 200
    payload = created.json()
    task_id = payload["task_id"]
    assert payload["status"] == "running"

    status = None
    for _ in range(80):
        response = client.get(f"/api/task/status/{task_id}")
        assert response.status_code == 200
        status = response.json()
        if status["status"] in {"completed", "failed"}:
            break
        time.sleep(0.05)

    assert status is not None
    assert status["status"] == "completed"
    assert status["progress"] == 100
    assert status["stage"]

    spec = client.get(f"/api/task/spec/{task_id}")
    assert spec.status_code == 200
    assert spec.json()["question_parse"]["disease"]
    assert spec.json()["study_design"]["research_type"]

    report = client.get(f"/api/task/report/{task_id}")
    assert report.status_code == 200
    assert report.json()["overall"] in {"PASS", "REVIEW", "REJECT"}
    assert len(report.json()["layers"]) == 4
    assert report.json()["cohort_f1"] is None


def _geo_matrix_result(tmp_path: Path, lines: list[str], *, accession: str = "GSE76360") -> GEOAdapterResult:
    matrix = tmp_path / f"{accession}_series_matrix.txt.gz"
    with gzip.open(matrix, "wt", encoding="utf-8") as handle:
        handle.write("\n".join(lines))
    source = SourceItem(
        source_id=f"geo:{accession}:test",
        task_id="task-geo",
        source_name="NCBI GEO",
        source_type="database",
        accession=accession,
        url="https://ftp.ncbi.nlm.nih.gov/test",
        file_type="series_matrix",
        local_path=str(matrix),
        checksum="sha256:test",
        status="downloaded",
    )
    return GEOAdapterResult(
        task_id="task-geo",
        accession=accession,
        portal_url=f"https://www.ncbi.nlm.nih.gov/geo/query/acc.cgi?acc={accession}",
        availability=[],
        resources=[
            GEOResourceRecord(
                accession=accession,
                resource_type=GEOResourceType.SERIES_MATRIX,
                file_name=matrix.name,
                download_url=source.url,
                status="downloaded",
                file_size=matrix.stat().st_size,
                source_item=source,
            )
        ],
        source_items=[source],
        cache_hit=GEOCacheStatus(accession_directory=False, resource_directories={}),
        queried_at=datetime.now(timezone.utc),
        notice="test",
    )


def _response_spec() -> ResearchSpec:
    return ResearchSpec(
        task_id="task-geo",
        research_goal=QUESTION,
        disease="Breast Cancer",
        subtype="HER2-positive",
        genes=["PIK3CA"],
        outcomes=["treatment_response"],
        required_data_types=["clinical", "treatment_response"],
    )


def test_geo_maps_synonym_and_concatenated_characteristics(tmp_path: Path) -> None:
    geo = _geo_matrix_result(
        tmp_path,
        [
            '!Sample_title\t"119_B"\t"120_B"',
            '!Sample_geo_accession\t"GSM1"\t"GSM3"',
            '!Sample_source_name_ch1\t"Breast tumor core biopsy"\t"Breast tumor core biopsy"',
            '!Sample_characteristics_ch1\t"subject id: 119; patient status: HER2+ Breast Cancer; timepoint: baseline; pathological complete response: pCR; er status: Pos"\t"subject id: 120; patient status: HER2+ Breast Cancer; timepoint: baseline; pathological complete response: NOR; er status: Neg"',
            "!series_matrix_table_begin",
        ],
    )

    built = ResearchDatasetBuilder().build_from_geo(geo, _response_spec())

    assert built is not None
    dataset, readiness = built
    assert dataset.row_count == 2
    assert dataset.target_column == "treatment_response" or dataset.target_column == "pcr"
    assert readiness.target_match is True
    assert {row["treatment_response"] for row in dataset.rows} == {"病理完全缓解（pCR）", "未达客观缓解"}
    assert all(row["sample_source"] == "Breast tumor core biopsy" for row in dataset.rows)
    assert all(row["sample_type"] == "原发肿瘤" for row in dataset.rows)
    assert all(row["sample_timepoint"] == "基线" for row in dataset.rows)
    assert all(row["disease"] == "乳腺癌" for row in dataset.rows)


def test_geo_response_cohort_does_not_report_zero_outcome_coverage(tmp_path: Path) -> None:
    geo = _geo_matrix_result(
        tmp_path,
        [
            '!Sample_title\t"119_B"\t"119_P"\t"120_B"\t"120_P"',
            '!Sample_geo_accession\t"GSM1"\t"GSM2"\t"GSM3"\t"GSM4"',
            '!Sample_source_name_ch1\t"Breast tumor core biopsy"\t"Breast tumor core biopsy"\t"Breast tumor core biopsy"\t"Breast tumor core biopsy"',
            '!Sample_characteristics_ch1\t"subject id: 119"\t"subject id: 119"\t"subject id: 120"\t"subject id: 120"',
            '!Sample_characteristics_ch1\t"patient status: HER2+ Breast Cancer"\t"patient status: HER2+ Breast Cancer"\t"patient status: HER2+ Breast Cancer"\t"patient status: HER2+ Breast Cancer"',
            '!Sample_characteristics_ch1\t"timepoint: baseline"\t"timepoint: post"\t"timepoint: baseline"\t"timepoint: post"',
            '!Sample_characteristics_ch1\t"response at surgery: pCR"\t"response at surgery: pCR"\t"response at surgery: NOR"\t"response at surgery: NOR"',
            "!series_matrix_table_begin",
        ],
    )
    spec = _response_spec()
    built = ResearchDatasetBuilder().build_from_geo(geo, spec)
    assert built is not None
    dataset, readiness = built
    design, _cohort = StudyDesignBuilder().build(spec, dataset, readiness, [], [])
    _iteration, critical, recommended = CollectionAgent().inspect(
        spec=spec,
        dataset=dataset,
        readiness=readiness,
        design=design,
        source_names=["NCBI GEO"],
        source_items=[],
        round_number=2,
        attempted_calls=set(),
        actions=[],
    )
    critical_ids = {gap.variable_id for gap in critical}
    by_id = {variable.variable_id: variable for variable in design.required_variables}

    assert by_id["outcome"].available is True
    assert by_id["treatment"].available is True
    assert by_id["disease"].available is True
    assert by_id["sample_type"].available is True
    assert by_id["sample_timepoint"].available is True
    assert by_id["sample_source"].available is True
    assert "outcome" not in critical_ids
    assert "treatment" not in critical_ids
    assert "disease" not in critical_ids
    assert "sample_type" not in critical_ids
    assert "sample_timepoint" not in critical_ids
    assert "sample_source" not in {gap.variable_id for gap in recommended}
    assert "pik3ca_mutation" in critical_ids


def test_same_patient_complete_pack_has_no_critical_gaps() -> None:
    spec = _response_spec()
    rows = [
        {
            "study_id": "breast_alpelisib_2020",
            "patient_id": f"P{index}",
            "sample_id": f"S{index}",
            "source_id": "cbioportal:breast_alpelisib_2020",
            "disease": "乳腺癌",
            "subtype": "HR-positive",
            "pik3ca_mutation": index % 2,
            "treatment": "Alpelisib",
            "treatment_response": "部分缓解" if index % 2 else "疾病进展",
            "sample_type": "原发肿瘤",
            "sample_timepoint": "治疗前",
        }
        for index in range(40)
    ]
    dataset = ResearchDatasetBuilder()._dataset_from_rows(
        rows,
        name="PIK3CA 突变乳腺癌 Alpelisib 治疗响应队列科研数据集",
        unit="患者",
        spec=spec,
    )
    readiness = ResearchDatasetBuilder()._readiness(dataset, spec)
    design, _cohort = StudyDesignBuilder().build(spec, dataset, readiness, [], [])
    _iteration, critical, recommended = CollectionAgent().inspect(
        spec=spec,
        dataset=dataset,
        readiness=readiness,
        design=design,
        source_names=["cBioPortal"],
        source_items=[],
        round_number=2,
        attempted_calls=set(),
        actions=[],
    )

    assert dataset.target_column == "treatment_response"
    assert readiness.requested_variable_coverage_rate == 1
    assert {gap.variable_id for gap in critical} == set()
    assert all(not gap.required for gap in recommended)


def test_geo_maps_pik3ca_mutation_from_sample_characteristics(tmp_path: Path) -> None:
    geo = _geo_matrix_result(
        tmp_path,
        [
            '!Sample_title\t"119_B"\t"120_B"',
            '!Sample_geo_accession\t"GSM1"\t"GSM3"',
            '!Sample_characteristics_ch1\t"subject id: 119"\t"subject id: 120"',
            '!Sample_characteristics_ch1\t"patient status: HER2+ Breast Cancer"\t"patient status: HER2+ Breast Cancer"',
            '!Sample_characteristics_ch1\t"timepoint: baseline"\t"timepoint: baseline"',
            '!Sample_characteristics_ch1\t"response at surgery: pCR"\t"response at surgery: NOR"',
            '!Sample_characteristics_ch1\t"pik3ca mutation: mutant"\t"pik3ca mutation: WT"',
            "!series_matrix_table_begin",
        ],
    )
    spec = _response_spec()
    built = ResearchDatasetBuilder().build_from_geo(geo, spec)
    assert built is not None
    dataset, readiness = built
    design, _cohort = StudyDesignBuilder().build(spec, dataset, readiness, [], [])
    _iteration, critical, _recommended = CollectionAgent().inspect(
        spec=spec,
        dataset=dataset,
        readiness=readiness,
        design=design,
        source_names=["NCBI GEO"],
        source_items=[],
        round_number=1,
        attempted_calls=set(),
        actions=[],
    )

    assert {row["pik3ca_mutation"] for row in dataset.rows} == {0, 1}
    assert readiness.requested_variable_coverage_rate == 1
    assert "pik3ca_mutation" not in {gap.variable_id for gap in critical}
    assert "outcome" not in {gap.variable_id for gap in critical}


def test_empty_geo_outcome_column_is_not_treated_as_matched_target(tmp_path: Path) -> None:
    geo = _geo_matrix_result(
        tmp_path,
        [
            '!Sample_title\t"s1"\t"s2"',
            '!Sample_geo_accession\t"GSM1"\t"GSM2"',
            '!Sample_characteristics_ch1\t"subject id: 1"\t"subject id: 2"',
            '!Sample_characteristics_ch1\t"patient status: HER2+ Breast Cancer"\t"patient status: HER2+ Breast Cancer"',
            "!series_matrix_table_begin",
        ],
        accession="GSE99999",
    )
    built = ResearchDatasetBuilder().build_from_geo(geo, _response_spec())
    assert built is not None
    dataset, readiness = built
    assert dataset.target_column is None
    assert readiness.target_match is False
    names = {column.name for column in dataset.columns}
    assert "treatment_response" not in names


def test_pcr_question_seeds_named_geo_when_source_budget_allows() -> None:
    question = "研究三阴性乳腺癌中 BRCA1/BRCA2 突变与新辅助化疗病理完全缓解（pCR）的关系，并整理患者级科研数据集"
    request = AgentTaskRequest(
        question=question,
        use_qwen=False,
        data_mode="plan_only",
        max_sources=8,
        max_records=200,
    )
    service = ResearchAgentService()
    spec = service._enrich_research_spec(service._deterministic_spec(question, "task-seed-pcr"), question)
    seeds = service._priority_seed_calls(spec, request)
    accessions = [
        str(call["arguments"].get("accession") or "")
        for call in seeds
        if call["name"] == "search_geo"
    ]
    merged = service._merge_tool_calls(seeds, service._deterministic_tool_calls(spec, request), request.max_sources)
    merged = service._prioritize_named_cohort_calls(merged, None)
    merged_geo = [
        str(call["arguments"].get("accession") or "")
        for call in merged
        if call["name"] == "search_geo"
    ]
    assert "GSE76360" in accessions
    assert "GSE25066" in accessions
    assert merged_geo[:2] == ["GSE76360", "GSE25066"]


def test_trial_question_seeds_ispy2_nct() -> None:
    question = "检索 I-SPY2 新辅助乳腺癌临床试验 NCT01042379 的登记信息"
    request = AgentTaskRequest(question=question, use_qwen=False, data_mode="plan_only", max_sources=8)
    service = ResearchAgentService()
    spec = service._enrich_research_spec(service._deterministic_spec(question, "task-seed-nct"), question)
    seeds = service._priority_seed_calls(spec, request)
    trial = next(call for call in seeds if call["name"] == "search_trials")
    assert trial["arguments"]["nct_id"] == "NCT01042379"


def test_cell_line_question_seeds_depmap() -> None:
    question = "整理乳腺癌细胞系对靶向药物的药敏（AUC/IC50），不得当作患者疗效。"
    request = AgentTaskRequest(question=question, use_qwen=False, data_mode="plan_only", max_sources=8)
    service = ResearchAgentService()
    spec = service._enrich_research_spec(service._deterministic_spec(question, "task-seed-depmap"), question)
    seeds = service._priority_seed_calls(spec, request)
    assert any(call["name"] == "search_depmap" for call in seeds)
    planned = service._deterministic_tool_calls(spec, request)
    assert any(call["name"] == "search_depmap" for call in planned)


def test_alpelisib_question_does_not_seed_her2_response_geo() -> None:
    question = "PIK3CA 突变的 HR+/HER2- 乳腺癌患者，使用阿培利司后的响应是否优于野生型？"
    request = AgentTaskRequest(question=question, use_qwen=False, data_mode="plan_only", max_sources=8)
    service = ResearchAgentService()
    spec = service._enrich_research_spec(service._deterministic_spec(question, "task-no-her2-geo"), question)
    seeds = service._priority_seed_calls(spec, request)
    accessions = {
        str(call["arguments"].get("accession") or "")
        for call in seeds
        if call["name"] == "search_geo"
    }
    assert "GSE76360" not in accessions
    assert "GSE25066" not in accessions


def _empty_dataset(name: str = "empty") -> ModelingDataset:
    return ModelingDataset(
        name=name,
        unit_of_analysis="患者",
        columns=[],
        rows=[],
        row_count=0,
        patient_count=0,
        sample_count=0,
    )


def _gap_readiness() -> AnalysisReadinessReport:
    return AnalysisReadinessReport(
        status="研究结局不匹配",
        analysis_ready=False,
        row_count=0,
        feature_count=0,
        split_strategy="按患者编号分组",
        target_match=False,
        requested_variable_coverage_rate=0.0,
    )


def test_pcr_gap_follow_up_emits_response_geo_queue() -> None:
    question = "研究三阴性乳腺癌中 BRCA1/BRCA2 突变与新辅助化疗病理完全缓解（pCR）的关系"
    service = ResearchAgentService()
    spec = service._enrich_research_spec(service._deterministic_spec(question, "task-follow-pcr"), question)
    request = AgentTaskRequest(question=question, use_qwen=False, data_mode="plan_only", max_sources=8)
    decision = service.collection_agent.decide(
        spec=spec,
        dataset=_empty_dataset("METABRIC"),
        readiness=_gap_readiness(),
        gaps=[
            CollectionGap(
                variable_id="outcome",
                label="研究结局",
                role="结局",
                required=True,
                coverage_rate=0.0,
                reason="当前主表无 pCR",
            )
        ],
        attempted_calls=set(),
        max_records=200,
        round_number=1,
        max_rounds=8,
    )
    calls = service._autonomous_follow_up_calls(
        spec=spec,
        request=request,
        decision=decision,
        raw_results=[],
        critical=[],
        dataset=_empty_dataset(),
        readiness=_gap_readiness(),
        attempted_calls=set(),
        qwen_client=None,
        round_number=1,
        max_rounds=8,
    )
    geo = [str(call["arguments"].get("accession") or "") for call in calls if call["name"] == "search_geo"]
    assert decision.action == "continue"
    assert any(accession in {"GSE76360", "GSE25066"} for accession in geo)


def test_cell_line_gap_follow_up_emits_depmap() -> None:
    question = "整理乳腺癌细胞系对靶向药物的药敏（AUC/IC50），不得当作患者疗效。"
    service = ResearchAgentService()
    spec = service._enrich_research_spec(service._deterministic_spec(question, "task-follow-depmap"), question)
    request = AgentTaskRequest(question=question, use_qwen=False, data_mode="plan_only", max_sources=8)
    decision = service.collection_agent.decide(
        spec=spec,
        dataset=_empty_dataset("none"),
        readiness=_gap_readiness(),
        gaps=[],
        attempted_calls=set(),
        max_records=80,
        round_number=1,
        max_rounds=8,
    )
    calls = service._autonomous_follow_up_calls(
        spec=spec,
        request=request,
        decision=decision,
        raw_results=[],
        critical=[],
        dataset=_empty_dataset(),
        readiness=_gap_readiness(),
        attempted_calls=set(),
        qwen_client=None,
        round_number=1,
        max_rounds=8,
    )
    assert any(call["name"] == "search_depmap" for call in calls)


def test_trial_gap_follow_up_emits_nct() -> None:
    question = "检索 I-SPY2 新辅助乳腺癌临床试验 NCT01042379 的登记信息"
    service = ResearchAgentService()
    spec = service._enrich_research_spec(service._deterministic_spec(question, "task-follow-nct"), question)
    request = AgentTaskRequest(question=question, use_qwen=False, data_mode="plan_only", max_sources=8)
    decision = service.collection_agent.decide(
        spec=spec,
        dataset=_empty_dataset("none"),
        readiness=_gap_readiness(),
        gaps=[
            CollectionGap(
                variable_id="evidence",
                label="试验证据",
                role="证据",
                required=True,
                coverage_rate=0.0,
                reason="尚未落到指定 NCT",
            )
        ],
        attempted_calls=set(),
        max_records=50,
        round_number=1,
        max_rounds=8,
    )
    calls = service._autonomous_follow_up_calls(
        spec=spec,
        request=request,
        decision=decision,
        raw_results=[],
        critical=[],
        dataset=_empty_dataset(),
        readiness=_gap_readiness(),
        attempted_calls=set(),
        qwen_client=None,
        round_number=1,
        max_rounds=8,
    )
    trials = [call for call in calls if call["name"] == "search_trials"]
    assert any(str(call["arguments"].get("nct_id") or "") == "NCT01042379" for call in trials)


def test_evidence_gap_follow_up_emits_paper_extract() -> None:
    question = "从开放论文表格与图注中整理乳腺癌 PIK3CA 与治疗响应的文献证据"
    service = ResearchAgentService()
    spec = service._enrich_research_spec(service._deterministic_spec(question, "task-follow-paper"), question)
    spec = spec.model_copy(update={"required_data_types": list(dict.fromkeys([*spec.required_data_types, "evidence"]))})
    request = AgentTaskRequest(question=question, use_qwen=False, data_mode="plan_only", max_sources=8)
    decision = service.collection_agent.decide(
        spec=spec,
        dataset=ModelingDataset(
            name="stub",
            unit_of_analysis="患者",
            columns=[],
            rows=[{}],
            row_count=1,
            patient_count=1,
            sample_count=1,
            target_column="treatment_response",
        ),
        readiness=AnalysisReadinessReport(
            status="解释层不足",
            analysis_ready=True,
            row_count=1,
            feature_count=1,
            split_strategy="按患者编号分组",
            target_match=True,
            requested_variable_coverage_rate=1.0,
        ),
        gaps=[
            CollectionGap(
                variable_id="evidence",
                label="文献证据",
                role="证据",
                required=True,
                coverage_rate=0.0,
                reason="尚未抽取论文表/图注",
            )
        ],
        attempted_calls=set(),
        max_records=20,
        round_number=1,
        max_rounds=8,
    )
    calls = service._autonomous_follow_up_calls(
        spec=spec,
        request=request,
        decision=decision,
        raw_results=[],
        critical=[],
        dataset=_empty_dataset(),
        readiness=_gap_readiness(),
        attempted_calls=set(),
        qwen_client=None,
        round_number=1,
        max_rounds=8,
    )
    assert any(call["name"] == "extract_paper_assets" for call in calls)


def test_qwen_seed_merge_respects_max_sources() -> None:
    question = "研究三阴性乳腺癌中 BRCA1/BRCA2 突变与新辅助化疗病理完全缓解（pCR）的关系"
    service = ResearchAgentService()
    spec = service._enrich_research_spec(service._deterministic_spec(question, "task-merge-cap"), question)
    request = AgentTaskRequest(question=question, use_qwen=False, max_sources=4)
    qwen_calls = [
        {"id": "qwen-1", "name": "search_cbioportal", "arguments": {"study_id": "brca_metabric"}},
        {"id": "qwen-2", "name": "search_geo_catalog", "arguments": {"query": "breast pCR"}},
    ]
    merged = service._merge_tool_calls(
        service._priority_seed_calls(spec, request) + qwen_calls,
        service._deterministic_tool_calls(spec, request),
        request.max_sources,
    )
    assert len(merged) <= 4
    accessions = [str(call["arguments"].get("accession") or "") for call in merged if call["name"] == "search_geo"]
    assert "GSE76360" in accessions or "GSE25066" in accessions
